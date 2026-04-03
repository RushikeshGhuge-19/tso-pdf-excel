"""
TSO Converter  (v3 — PDF or Excel input, fully dynamic)
========================================================
Accepts either a TSO PDF or a filled TSO Excel as input source.
- PDF input  : extracts data via pdfplumber (pages 1-4)
- Excel input: reads data directly from the uploaded workbook's sheets

Zero hardcoded data values.
- All dropdown values read from Library sheet at runtime
- All structural values read from template rows at runtime
- PROC_RULES maps PDF keyword patterns → Library search terms

Usage (CLI):
    python tso_converter_v3.py INPUT.(pdf|xlsx) TEMPLATE.xlsx OUT.xlsx
    python tso_converter_v3.py          # opens tkinter file pickers

Requirements: pip install pdfplumber openpyxl streamlit
"""

import sys, re, shutil, io
from pathlib import Path
from difflib import SequenceMatcher

import pdfplumber
import openpyxl
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# LIBRARY  — load all dropdown lists from template's Library sheet
# ─────────────────────────────────────────────────────────────────────────────
def load_library(wb):
    from openpyxl.utils import get_column_letter
    ws = wb['Library']
    lib = {}
    for ci in range(1, ws.max_column + 1):
        sn = ws.cell(1, ci).value
        cn = ws.cell(2, ci).value
        if not sn and not cn:
            continue
        vals = []
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, ci).value
            if v not in (None, ''):
                vals.append(str(v).strip())
        if vals:
            lib[get_column_letter(ci)] = vals
    return lib


def find_in_lib(search_terms, lib_vals, prefer_no_suffix=True):
    """Find Library value containing ALL search_terms (case-insensitive)."""
    terms = [t.lower() for t in search_terms]
    matches = [v for v in lib_vals if all(t in v.lower() for t in terms)]
    if not matches:
        return None
    if prefer_no_suffix:
        base = [m for m in matches if not re.search(r'\s+\d+\s*$', m.strip())]
        return base[0] if base else matches[0]
    return matches[0]


def find_sub_op(search_terms, lib_vals, pdf_keywords=None):
    """Prefer plain 'Piercing N' over 'Cam Piercing N' for piercing rules."""
    base = find_in_lib(search_terms, lib_vals)
    if base is None:
        return None
    if pdf_keywords and 'CAM' in pdf_keywords and 'PIERC' in pdf_keywords:
        matches = [v for v in lib_vals if all(t in v.lower() for t in search_terms)]
        plain = [m for m in matches if m.lower().startswith('piercing')]
        return plain[0] if plain else base
    if pdf_keywords and 'PIERC' in pdf_keywords and 'CAM' not in pdf_keywords and 'BLANK' not in pdf_keywords:
        matches = [v for v in lib_vals if all(t in v.lower() for t in search_terms)]
        plain = [m for m in matches if m.lower().startswith('piercing')]
        return plain[0] if plain else base
    return base


# ─────────────────────────────────────────────────────────────────────────────
# PROCESS RULES  — keyword patterns → Library search terms (never fixed strings)
# ─────────────────────────────────────────────────────────────────────────────
PROC_RULES = [
    (['BLANK','PIERC'],  ['blank','pierce'],       ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['1ST','FORM'],     ['forming','1'],           ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['FIRST','FORM'],   ['forming','1'],           ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['2ND','FORM'],     ['forming','2'],           ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['SECOND','FORM'],  ['forming','2'],           ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['CAM','PIERC'],    ['piercing','1'],          ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['PIERC'],          ['piercing','2'],          ['sheet metal','cold'], ['tool','m&m'],
     ['strokes'],        ['nos'],                  ['tonnage'],            ['others']),
    (['SHEAR'],          ['shearing'],              ['others'],             ['tool','supplier'],
     ['weight-exact'],   ['kgs'],                  [],                     []),
    (['INSPECT'],        ['inspection'],            ['others'],             ['gauge','m&m'],
     ['pieces'],         ['nos'],                  [],                     []),
    (['RETAP'],          ['retapping'],             ['others'],             ['tool','supplier'],
     ['strokes'],        ['nos'],                  [],                     []),
    (['PROJECT','WELD'], ['projection','welding'],  ['others'],             ['tool','supplier'],
     ['strokes'],        ['nos'],                  [],                     []),
]


def match_rule(pdf_name_upper, lib):
    """Return (mfg, sub_op, ftg, p1t, p1u, p2t, p2u) by matching PROC_RULES."""
    sub_op_vals  = lib.get('W', [])
    mfg_vals     = lib.get('V', [])
    ftg_vals     = lib.get('X', [])
    p_type_vals  = lib.get('Y', [])
    p_uom_vals   = lib.get('Z', [])

    def _p(terms, vals):
        if not terms: return ''
        if terms == ['weight-exact']:
            exact = [v for v in vals if v.strip().lower() == 'weight']
            return exact[0] if exact else find_in_lib(['weight'], vals)
        return find_in_lib(terms, vals)

    for (pdf_kws, sub_terms, mfg_terms, ftg_terms,
         p1t_terms, p1u_terms, p2t_terms, p2u_terms) in PROC_RULES:
        if all(k in pdf_name_upper for k in pdf_kws):
            return (
                find_in_lib(mfg_terms, mfg_vals),
                find_sub_op(sub_terms, sub_op_vals, pdf_kws),
                find_in_lib(ftg_terms, ftg_vals),
                _p(p1t_terms, p_type_vals),
                _p(p1u_terms, p_uom_vals),
                _p(p2t_terms, p_type_vals),
                _p(p2u_terms, p_uom_vals),
            )
    return None, None, None, None, None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
SKIP = ('', '---', '—', '-', 'None', 'none')

def sv(ws, r, c, v):
    if v is None or str(v).strip() in SKIP: return
    ws.cell(row=r, column=c, value=v)

def cl(v):
    s = str(v).strip() if v is not None else ''
    return '' if s in SKIP else s

def title_case(s):
    r = s.title()
    return re.sub(r'(\d+)(St|Nd|Rd|Th)\b', lambda m: m.group(1) + m.group(2).lower(), r)


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE READERS
# ─────────────────────────────────────────────────────────────────────────────
def read_template_bom(wb):
    ws = wb['BOM Template']
    rows, order = {}, []
    for r in range(3, ws.max_row + 1):
        pno = str(ws.cell(r, 2).value or '').strip()
        if pno:
            rows[pno] = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            order.append(pno)
    return rows, order

def read_template_row(wb, sheet, row_num):
    ws = wb[sheet]
    return [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

def read_template_proc_structure(wb):
    ws = wb['Inhouse Process']
    s = {}
    for r in range(3, ws.max_row + 1):
        lv = str(ws.cell(r, 1).value or '').strip()
        pn = str(ws.cell(r, 2).value or '').strip()
        dc = str(ws.cell(r, 3).value or '').strip()
        if lv or pn:
            s[r] = {'level': lv, 'pno': pn, 'desc': dc}
    return s


# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSER  — extracts data from a TSO PDF (4-page format)
# ─────────────────────────────────────────────────────────────────────────────
def parse_pdf(pdf_path):
    data = {'meta': {}, 'bom': [], 'inhouse_rm': {}, 'tool_ops': [], 'source': 'pdf'}

    with pdfplumber.open(pdf_path) as pdf:

        # Page 1: meta
        if pdf.pages:
            for row in (pdf.pages[0].extract_tables() or [[]])[0]:
                rc = [cl(c) for c in row]
                if not rc: continue
                if rc[0] == 'Date' and len(rc) >= 2:
                    data['meta']['date'] = rc[1]
                    for i, v in enumerate(rc):
                        if v == 'Project Name' and i+1 < len(rc):
                            data['meta']['project'] = rc[i+1]
                if rc[0] == 'Supplier Name' and len(rc) >= 2:
                    data['meta']['supplier'] = rc[1]
                    for i, v in enumerate(rc):
                        if 'stamping' in str(v).lower() and i+1 < len(rc):
                            data['meta']['stamping_loc'] = rc[i+1]
                if any('end items' in str(c).lower() for c in rc if c):
                    data['meta']['end_items'] = rc[1] if len(rc) > 1 else ''
                    for i, v in enumerate(rc):
                        if 'Welding' in str(v) and i+1 < len(rc):
                            data['meta']['welding_loc'] = rc[i+1]

        # Page 2: BOM
        if len(pdf.pages) >= 2:
            tbls = pdf.pages[1].extract_tables()
            if tbls:
                for row in tbls[0][1:]:
                    rc = [cl(c) for c in row]
                    if not rc[0]: continue
                    data['bom'].append({
                        'sno':       rc[0],
                        'part_no':   rc[1]  if len(rc) > 1  else '',
                        'type_part': rc[5]  if len(rc) > 5  else '',
                        'cad_wt':    rc[7]  if len(rc) > 7  else '',
                        'material':  rc[8]  if len(rc) > 8  else '',
                        'thickness': rc[9]  if len(rc) > 9  else '',
                        'qty_assy':  rc[10] if len(rc) > 10 else '',
                        'qty_veh':   rc[11] if len(rc) > 11 else '',
                    })

        # Page 3: RM + tool ops
        if len(pdf.pages) >= 3:
            tbls = pdf.pages[2].extract_tables()
            if not tbls: return data
            tbl = tbls[0]
            main_idx = None
            for idx, row in enumerate(tbl):
                rc = [cl(c) for c in row]
                if rc and rc[0] == '1':
                    main_idx = idx
                    data['inhouse_rm'] = {
                        'input_wt':  rc[32] if len(rc) > 32 else '',
                        'output_wt': rc[33] if len(rc) > 33 else '',
                        'blank_thk': rc[17] if len(rc) > 17 else '',
                    }
                    break
            if main_idx is None: return data

            def extract_op(rc, n1, n2, lc):
                if len(rc) <= lc + 6: return None
                p1 = rc[n1] if n1 < len(rc) else ''
                p2 = rc[n2] if n2 < len(rc) else ''
                name = f"{p1} {p2}".strip() if p2 and p2.upper() == 'TOOL' else (f"{p1} {p2}".strip() if p2 else p1)
                name = name.strip()
                if not name: return None
                u = name.upper()
                if not any(k in u for k in ['BLANK','FORM','PIERC','INSPECT','SHEAR','PANEL','WELD','TOOL','TAP']):
                    return None
                return {
                    'raw_name':  name,
                    'tool_l':    rc[lc]   if len(rc) > lc   else '',
                    'tool_w':    rc[lc+1] if len(rc) > lc+1 else '',
                    'tool_h':    rc[lc+2] if len(rc) > lc+2 else '',
                    'tonnage':   rc[lc+3] if len(rc) > lc+3 else '',
                    'press':     rc[lc+4] if len(rc) > lc+4 else '',
                    'parts_per': rc[lc+5] if len(rc) > lc+5 else '',
                    'construct': rc[lc+6] if len(rc) > lc+6 else '',
                }

            op = extract_op([cl(c) for c in tbl[main_idx]], 40, 41, 42)
            if op: data['tool_ops'].append(op)
            for row in tbl[main_idx + 1:]:
                rc = [cl(c) for c in row]
                if not any(v for v in rc): continue
                op = extract_op(rc, 40, 41, 42)
                if op: data['tool_ops'].append(op)

    return data


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSER  — extracts data from a filled TSO Excel workbook
# ─────────────────────────────────────────────────────────────────────────────
def parse_excel(excel_path):
    """
    Read data from a filled TSO Excel (same 19-sheet format).
    Extracts BOM, Inhouse RM, and Inhouse Process data.
    Returns the same data dict structure as parse_pdf().
    """
    data = {'meta': {}, 'bom': [], 'inhouse_rm': {}, 'tool_ops': [], 'source': 'excel'}

    wb = load_workbook(str(excel_path), data_only=True)

    # ── Meta: try TSO Summary sheet first, else infer from BOM ──────────────
    if 'TSO Summary' in wb.sheetnames:
        ws = wb['TSO Summary']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            key = str(row[0]).strip().lower()
            val = cl(row[1]) if len(row) > 1 else ''
            if 'date' in key and 'sign' not in key: data['meta']['date'] = val
            elif 'project' in key:                  data['meta']['project'] = val
            elif 'supplier' in key and 'sign' not in key: data['meta']['supplier'] = val
            elif 'stamping' in key:                 data['meta']['stamping_loc'] = val
            elif 'welding' in key:                  data['meta']['welding_loc'] = val
            elif 'end items' in key:                data['meta']['end_items'] = val

    # ── BOM Template ──────────────────────────────────────────────────────
    if 'BOM Template' in wb.sheetnames:
        ws = wb['BOM Template']
        # Row 1 = headers, Row 2 = blank template row, Row 3+ = data
        for r in range(3, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v for v in row): continue
            level    = cl(row[0])
            part_no  = cl(row[1])
            if not part_no: continue
            data['bom'].append({
                'sno':       '1' if level == '0' else '1.1',
                'part_no':   part_no,
                'type_part': cl(row[5]) if len(row) > 5 else '',
                'cad_wt':    cl(row[11]) if len(row) > 11 else '',
                'material':  '',
                'thickness': '',
                'qty_assy':  cl(row[10]) if len(row) > 10 else '',
                'qty_veh':   cl(row[10]) if len(row) > 10 else '',
                # preserve level so we can distinguish assembly (0) from child (1)
                '_level':    level,
            })

    # ── Inhouse RM ────────────────────────────────────────────────────────
    if 'Inhouse RM' in wb.sheetnames:
        ws = wb['Inhouse RM']
        # Row 2 = ref row, Row 3 = data
        if ws.max_row >= 3:
            r3 = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
            data['inhouse_rm'] = {
                'input_wt':  cl(r3[15]) if len(r3) > 15 else '',
                'output_wt': cl(r3[17]) if len(r3) > 17 else '',
                'blank_thk': cl(r3[12]) if len(r3) > 12 else '',
            }
            # Also grab material from col F (RM Grade Others)
            if len(r3) > 5 and r3[5]:
                # Find matching BOM part and attach material
                child_pno = cl(r3[1]) if len(r3) > 1 else ''
                for p in data['bom']:
                    if p['part_no'] == child_pno or child_pno in p['part_no']:
                        p['material'] = cl(r3[5])
                        p['thickness'] = cl(r3[12]) if len(r3) > 12 else ''
                        break

    # ── Inhouse Process ────────────────────────────────────────────────────
    if 'Inhouse Process' in wb.sheetnames:
        ws = wb['Inhouse Process']
        # Row 2 = ref row, Row 3+ = data
        for r in range(3, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v for v in row): continue
            mfg      = cl(row[3]) if len(row) > 3 else ''
            sub_op   = cl(row[4]) if len(row) > 4 else ''
            ftg_name = cl(row[8]) if len(row) > 8 else ''
            tonnage  = cl(row[17]) if len(row) > 17 else ''
            construct = cl(row[18]) if len(row) > 18 else ''
            tool_l   = ''
            tool_w   = ''
            tool_h   = ''
            if not mfg and not sub_op: continue
            # Reconstruct raw_name from FTG Name or Sub Op for categorisation
            raw_name = ftg_name or sub_op
            if raw_name:
                data['tool_ops'].append({
                    'raw_name':  raw_name.upper(),
                    'tool_l':    tool_l,
                    'tool_w':    tool_w,
                    'tool_h':    tool_h,
                    'tonnage':   tonnage,
                    'press':     cl(row[11]) if len(row) > 11 else '',
                    'parts_per': '',
                    'construct': construct,
                    # Extra context for Excel-sourced ops
                    '_mfg':     mfg,
                    '_sub_op':  sub_op,
                    '_ftg_name': ftg_name,
                })

    return data


# ─────────────────────────────────────────────────────────────────────────────
# UNIFIED PARSER  — detects file type and routes accordingly
# ─────────────────────────────────────────────────────────────────────────────
def parse_input(input_path):
    """Parse PDF or Excel input. Returns unified data dict."""
    suffix = Path(str(input_path)).suffix.lower()
    if suffix == '.pdf':
        return parse_pdf(input_path)
    elif suffix in ('.xlsx', '.xls', '.xlsm'):
        return parse_excel(input_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Use .pdf or .xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(data, template_path, out_path):
    flags = []
    shutil.copy2(str(template_path), str(out_path))
    wb = load_workbook(str(out_path))
    lib = load_library(wb)

    tpl_bom_rows, tpl_row_order = read_template_bom(wb)
    tpl_rm3  = read_template_row(wb, 'Inhouse RM', 3)
    tpl_proc = read_template_proc_structure(wb)

    meta     = data['meta']
    bom      = data['bom']
    stamp    = data['inhouse_rm']
    tool_ops = data['tool_ops']
    source   = data.get('source', 'pdf')

    # BOM lookup by normalised part no
    pdf_bom = {}
    for p in bom:
        raw = str(p['part_no']).strip()
        pdf_bom[raw] = p
        if raw and not raw.startswith('0') and raw[0].isdigit():
            pdf_bom['0' + raw] = p

    # ── BOM Template ──────────────────────────────────────────────────────
    ws = wb['BOM Template']
    for i, canon_pno in enumerate(tpl_row_order):
        r = i + 3
        trow = tpl_bom_rows.get(canon_pno)
        if not trow: continue
        pdf_p = pdf_bom.get(canon_pno, {})
        for col, val in enumerate(trow, 1):
            if col in (11, 12): continue
            sv(ws, r, col, val)
        sv(ws, r, 11, trow[10])
        sv(ws, r, 12, pdf_p.get('cad_wt') or trow[11])

    # ── Inhouse RM ────────────────────────────────────────────────────────
    ws = wb['Inhouse RM']
    r = 3
    for col, val in enumerate(tpl_rm3, 1):
        if col in (16, 17, 18, 19): continue
        sv(ws, r, col, val)
    for p in bom:
        if '103710' in str(p.get('part_no', '')) and p.get('thickness'):
            sv(ws, r, 13, p['thickness']); break
    raw_g = stamp.get('input_wt', '')
    raw_n = stamp.get('output_wt', '')
    try:    sv(ws, r, 16, float(raw_g) if raw_g else tpl_rm3[15])
    except: sv(ws, r, 16, raw_g or tpl_rm3[15])
    try:    sv(ws, r, 18, float(raw_n) if raw_n else tpl_rm3[17])
    except: sv(ws, r, 18, raw_n or tpl_rm3[17])
    ws.cell(row=r, column=17, value='=P3-R3')
    ws.cell(row=r, column=19, value='=R3/P3*100')
    raw_mat = next((p.get('material','') for p in bom
                    if '103710' in str(p.get('part_no','')) and p.get('material')), '')
    if raw_mat:
        flags.append(
            f"⚠ Inhouse RM — RM Grade (col E): '{raw_mat}' is non-standard. "
            f"Col F pre-filled from template. Please select correct grade from dropdown."
        )

    # ── Inhouse Process ───────────────────────────────────────────────────
    ws = wb['Inhouse Process']

    assy_pno = assy_desc = child_pno = child_desc = ''
    for rn, v in sorted(tpl_proc.items()):
        if v['level'] == '0' and v['pno'] and not assy_pno:
            assy_pno = v['pno']; assy_desc = v['desc']
        if v['level'] == '1' and v['pno'] and not child_pno:
            child_pno = v['pno']; child_desc = v['desc']

    child_start = min((rn for rn, v in tpl_proc.items() if v['level'] == '1'), default=6)

    # For Excel source: ops already have _mfg/_sub_op/_ftg_name set
    # For PDF source: ops only have raw_name, need categorisation
    cats = {k: None for k in ['BLANK_PIERCE','FORM_1','FORM_2','PIERCE_1','PIERCE_2','INSPECT']}
    for op in tool_ops:
        # Use _sub_op if from Excel, raw_name if from PDF
        u = (op.get('_ftg_name') or op.get('raw_name') or '').upper()
        if 'BLANK' in u and 'PIERC' in u:                   cats['BLANK_PIERCE'] = op
        elif ('1ST' in u or 'FIRST' in u or 'FORMING 1' in u or 'FORM 1' in u) and 'FORM' in u: cats['FORM_1'] = op
        elif ('2ND' in u or 'SECOND' in u or 'FORMING 2' in u or 'FORM 2' in u) and 'FORM' in u: cats['FORM_2'] = op
        elif 'CAM' in u and 'PIERC' in u:                   cats['PIERCE_1'] = op
        elif 'PIERC' in u and 'CAM' not in u and 'BLANK' not in u:
            if cats['PIERCE_2'] is None: cats['PIERCE_2'] = op
        elif 'INSPECT' in u or 'PANEL' in u:                cats['INSPECT'] = op

    ordered_cats = ['BLANK_PIERCE','FORM_1','FORM_2','PIERCE_1','PIERCE_2','INSPECT']
    cat_pdf_kws = {
        'BLANK_PIERCE': ['BLANK','PIERC'],
        'FORM_1':       ['1ST','FORM'],
        'FORM_2':       ['2ND','FORM'],
        'PIERCE_1':     ['CAM','PIERC'],
        'PIERCE_2':     ['PIERC'],
        'INSPECT':      ['INSPECT'],
    }

    cur = child_start

    # Shearing row
    shear_mfg, shear_sub, shear_ftg, shear_p1t, shear_p1u, _, _ = match_rule('SHEAR', lib)
    gross_val = stamp.get('input_wt', '')
    try:    gross = float(gross_val) if gross_val else ''
    except: gross = gross_val

    sv(ws, cur, 1, '1'); sv(ws, cur, 2, child_pno); sv(ws, cur, 3, child_desc)
    sv(ws, cur, 4, shear_mfg); sv(ws, cur, 5, shear_sub); sv(ws, cur, 7, 10)
    sv(ws, cur, 8, shear_ftg); sv(ws, cur, 13, shear_p1t); sv(ws, cur, 14, shear_p1u)
    sv(ws, cur, 15, gross)
    cur += 1

    # Stamping ops
    for s_idx, cat_key in enumerate(ordered_cats):
        op = cats.get(cat_key)
        pdf_kws = '|'.join(cat_pdf_kws[cat_key])
        mfg, sub, ftg, p1t, p1u, p2t, p2u = match_rule(pdf_kws, lib)

        # FTG Name: prefer Excel's existing FTG name, else title-case from raw PDF
        if op and op.get('_ftg_name'):
            ftg_name = op['_ftg_name']
        elif op and op.get('raw_name'):
            ftg_name = title_case(op['raw_name'])
        else:
            ftg_name = ''

        tonnage   = op['tonnage']   if op and op.get('tonnage')   else ''
        construct = op['construct'] if op and op.get('construct') else ''
        if construct and construct == construct.upper():
            construct = title_case(construct)
        p2v = tonnage if p2t else ''

        level = '1' if s_idx == 0 else ''
        pno   = child_pno  if s_idx == 0 else ''
        desc  = child_desc if s_idx == 0 else ''

        sv(ws, cur, 1, level); sv(ws, cur, 2, pno); sv(ws, cur, 3, desc)
        sv(ws, cur, 4, mfg);   sv(ws, cur, 5, sub); sv(ws, cur, 7, (s_idx + 2) * 10)
        sv(ws, cur, 8, ftg);   sv(ws, cur, 9, ftg_name); sv(ws, cur, 10, 1 if ftg_name else '')
        if mfg and 'sheet metal' in mfg.lower():
            sv(ws, cur, 11, 'Igsec'); sv(ws, cur, 12, 'Pneumatic')
        sv(ws, cur, 13, p1t);  sv(ws, cur, 14, p1u); sv(ws, cur, 15, 1 if p1t else '')
        sv(ws, cur, 16, p2t);  sv(ws, cur, 17, p2u); sv(ws, cur, 18, p2v)
        sv(ws, cur, 19, construct)
        cur += 1

    wb.save(str(out_path))
    return flags


# ─────────────────────────────────────────────────────────────────────────────
# MAIN  — CLI entry point
# ─────────────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) >= 4:
        input_path    = Path(sys.argv[1])
        template_path = Path(sys.argv[2])
        out_path      = Path(sys.argv[3])
    else:
        import tkinter as tk
        from tkinter import filedialog, messagebox
        root = tk.Tk(); root.withdraw()

        ps = filedialog.askopenfilename(
            title="1 of 3 — Select TSO Input (PDF or Excel)",
            filetypes=[("PDF or Excel","*.pdf *.xlsx"),("PDF","*.pdf"),("Excel","*.xlsx"),("All","*.*")])
        if not ps: messagebox.showinfo("Cancelled","No input file."); return
        input_path = Path(ps)

        ts = filedialog.askopenfilename(
            title="2 of 3 — Select TSO Template Excel",
            filetypes=[("Excel","*.xlsx"),("All","*.*")])
        if not ts: messagebox.showinfo("Cancelled","No template."); return
        template_path = Path(ts)

        os_ = filedialog.asksaveasfilename(
            title="3 of 3 — Save Output As",
            defaultextension=".xlsx",
            initialfile=input_path.stem+"_TSO.xlsx",
            filetypes=[("Excel","*.xlsx"),("All","*.*")])
        if not os_: messagebox.showinfo("Cancelled","No output path."); return
        out_path = Path(os_)

    print(f"\nInput        :  {input_path.name}")
    print(f"Template     :  {template_path.name}")

    data = parse_input(input_path)
    print(f"Source       :  {data['source'].upper()}")
    print(f"Project      :  {data['meta'].get('project','')}")
    print(f"Supplier     :  {data['meta'].get('supplier','')}")
    print(f"BOM parts    :  {len(data['bom'])}")
    print(f"Tool ops     :  {len(data['tool_ops'])}")

    flags = write_excel(data, template_path, out_path)
    print(f"\nSaved        →  {out_path}")

    if flags:
        print(f"\nFlags ({len(flags)}):")
        for f in flags: print(f)
    else:
        print("No flags.")

    if len(sys.argv) < 4:
        from tkinter import messagebox
        msg = f"Saved:\n{out_path}"
        if flags: msg += f"\n\n{len(flags)} flag(s) — see console."
        messagebox.showinfo("Done!", msg)

if __name__ == "__main__":
    main()

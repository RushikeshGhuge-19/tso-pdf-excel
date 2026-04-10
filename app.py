"""
TSO Converter  (v4 — PDF or Excel input, fully dynamic)
========================================================
Accepts either a TSO PDF or a filled TSO Excel as input source.
- PDF input  : extracts data via pdfplumber (pages 1-4)
- Excel input: reads data directly from the uploaded workbook's sheets


Usage (CLI):
    python app.py INPUT.(pdf|xlsx) TEMPLATE.xlsx OUT.xlsx
    python app.py          # opens tkinter file pickers

Requirements: pip install pdfplumber openpyxl streamlit
"""

import sys, re, shutil, io, copy
from pathlib import Path
from difflib import SequenceMatcher

import pdfplumber
import openpyxl
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# LIBRARY  — load all dropdown lists from template's Library sheet
# ─────────────────────────────────────────────────────────────────────────────
def load_library(wb):
    from openpyxl.utils import get_column_letter
    ws = wb['Library']
    lib = {}
    for ci in range(1, ws.max_column + 1):
        sn = ws.cell(1, ci).value
        cn = ws.cell(2, ci).value
        if not sn and not cn:
            continue
        vals = []
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, ci).value
            if v not in (None, ''):
                vals.append(str(v).strip())
        if vals:
            lib[get_column_letter(ci)] = vals
    return lib


def find_in_lib(search_terms, lib_vals, prefer_no_suffix=True):
    """Find Library value containing ALL search_terms (case-insensitive).
    When prefer_no_suffix=True, prefer shorter plain values over numbered variants.
    Also prefer values that start with the first search term (avoids false-prefix matches
    like 'Arm stamping,bending, forming' for ['forming']).
    """
    terms = [t.lower() for t in search_terms]
    matches = [v for v in lib_vals if all(t in v.lower() for t in terms)]
    if not matches:
        return None
    if prefer_no_suffix:
        base = [m for m in matches if not re.search(r'\s+\d+\s*$', m.strip())]
        candidates = base if base else matches
        # Prefer values that START with the primary search term (e.g. 'Forming' over 'Arm stamping...')
        first_term = terms[0]
        starts_with = [m for m in candidates if m.strip().lower().startswith(first_term)]
        if starts_with:
            return starts_with[0]
        return candidates[0]
    return matches[0]


def find_sub_op(search_terms, lib_vals, pdf_keywords=None):
    """Prefer plain 'Piercing N' over 'Cam Piercing N' for piercing/punching rules."""
    base = find_in_lib(search_terms, lib_vals)
    if base is None:
        return None
    is_pierc_type = pdf_keywords and (
        'PIERC' in pdf_keywords or 'PUNCH' in pdf_keywords
    ) and 'BLANK' not in pdf_keywords
    if is_pierc_type:
        kws = pdf_keywords if not isinstance(pdf_keywords, list) else '|'.join(pdf_keywords)
        matches = [v for v in lib_vals if all(t in v.lower() for t in
                   [s.lower() for s in search_terms])]
        plain = [m for m in matches if m.lower().startswith('piercing')]
        return plain[0] if plain else base
    return base


# ─────────────────────────────────────────────────────────────────────────────
# PROCESS RULES  — keyword patterns → Library search terms (never fixed strings)
# ─────────────────────────────────────────────────────────────────────────────
PROC_RULES = [
    (['BLANK','PIERC'],  ['blank','pierce'],        ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['1ST','FORM'],     ['forming','1'],           ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['FIRST','FORM'],   ['forming','1'],           ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['2ND','FORM'],     ['forming','2'],           ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['SECOND','FORM'],  ['forming','2'],           ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['FORM'],           ['forming'],               ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['CAM','PIERC'],    ['piercing','1'],          ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['1ST','PUNCH'],    ['piercing','1'],          ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['FIRST','PUNCH'],  ['piercing','1'],          ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['2ND','PUNCH'],    ['piercing','2'],          ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['SECOND','PUNCH'], ['piercing','2'],          ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['PIERC'],          ['piercing','2'],          ['sheet metal','cold'], ['tool','m&m'],
     ['parts/stroke'],   ['nos'],                  ['tonnage'],            ['others']),
    (['SHEAR'],          ['shearing'],              ['others'],             ['tool','supplier'],
     ['weight-exact'],   ['kgs'],                  [],                     []),
    (['INSPECT'],        ['inspection'],            ['others'],             ['gauge','m&m'],
     ['others'],         ['nos'],                  [],                     []),
    (['RETAP'],          ['retapping'],             ['others'],             ['tool','supplier'],
     ['parts/stroke'],   ['nos'],                  [],                     []),
    (['PROJECT','WELD'], ['projection','welding'],  ['others'],             ['tool','supplier'],
     ['parts/stroke'],   ['nos'],                  [],                     []),
    (['RIVET'],          ['riveting'],              ['others'],             ['fixture','m&m'],
     ['parts/stroke'],   ['nos'],                  [],                     []),
]


def match_rule(pdf_name_upper, lib):
    """Return (mfg, sub_op, ftg, p1t, p1u, p2t, p2u) by matching PROC_RULES."""
    sub_op_vals  = lib.get('W', [])
    mfg_vals     = lib.get('V', [])
    ftg_vals     = lib.get('X', [])
    p_type_vals  = lib.get('Y', [])
    p_uom_vals   = lib.get('Z', [])

    def _p(terms, vals):
        if not terms: return ''
        if terms == ['weight-exact']:
            exact = [v for v in vals if v.strip().lower() == 'weight']
            return exact[0] if exact else find_in_lib(['weight'], vals)
        return find_in_lib(terms, vals)

    for (pdf_kws, sub_terms, mfg_terms, ftg_terms,
         p1t_terms, p1u_terms, p2t_terms, p2u_terms) in PROC_RULES:
        if all(k in pdf_name_upper for k in pdf_kws):
            return (
                find_in_lib(mfg_terms, mfg_vals),
                find_sub_op(sub_terms, sub_op_vals, pdf_kws),
                find_in_lib(ftg_terms, ftg_vals),
                _p(p1t_terms, p_type_vals),
                _p(p1u_terms, p_uom_vals),
                _p(p2t_terms, p_type_vals),
                _p(p2u_terms, p_uom_vals),
            )
    return None, None, None, None, None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
SKIP = ('', '---', '—', '-', 'None', 'none')

def sv(ws, r, c, v):
    if v is None or str(v).strip() in SKIP: return
    ws.cell(row=r, column=c, value=v)

def cl(v):
    s = str(v).strip() if v is not None else ''
    return '' if s in SKIP else s

def title_case(s):
    r = s.title()
    return re.sub(r'(\d+)(St|Nd|Rd|Th)\b', lambda m: m.group(1) + m.group(2).lower(), r)

def normalise_ftg_name(raw):
    """Clean up raw PDF tool name → proper FTG Name."""
    if not raw:
        return ''
    u = raw.upper().strip()
    if 'INSPECT' in u or 'PANEL' in u or 'CHECK' in u:
        side = ''
        if u.endswith('RH') or ' RH' in u: side = ' RH'
        elif u.endswith('LH') or ' LH' in u: side = ' LH'
        return f'Panel checker{side}'
    tc = title_case(raw)
    tc = re.sub(r'\bForm Tool\b', 'Forming Tool', tc)
    tc = re.sub(r'\b1st Punching.*Tool\b', '1st Piercing Tool', tc, flags=re.IGNORECASE)
    tc = re.sub(r'\b2nd Punching.*Tool\b', '2nd Piercing Tool', tc, flags=re.IGNORECASE)
    if tc.strip().lower() == 'piercing':
        return 'Piercing Tool'
    return tc


def select_child_part(bom):
    """Return the first likely inhouse child part from BOM rows."""
    for p in bom:
        level = str(p.get('_level', '')).strip()
        if level == '1' and p.get('part_no'):
            return p
    for p in bom:
        tp = str(p.get('type_part', '')).upper()
        if tp not in ('BOU', 'BOP') and p.get('sno', '') != '1' and p.get('part_no'):
            return p
    return None


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE READERS
# ─────────────────────────────────────────────────────────────────────────────
def read_template_bom(wb):
    ws = wb['BOM Template']
    rows, order = {}, []
    for r in range(3, ws.max_row + 1):
        pno = str(ws.cell(r, 2).value or '').strip()
        if pno:
            rows[pno] = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            order.append(pno)
    return rows, order

def read_template_row(wb, sheet, row_num):
    ws = wb[sheet]
    return [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

def read_template_proc_structure(wb):
    ws = wb['Inhouse Process']
    s = {}
    for r in range(3, ws.max_row + 1):
        lv = str(ws.cell(r, 1).value or '').strip()
        pn = str(ws.cell(r, 2).value or '').strip()
        dc = str(ws.cell(r, 3).value or '').strip()
        if lv or pn:
            s[r] = {'level': lv, 'pno': pn, 'desc': dc}
    return s


# ─────────────────────────────────────────────────────────────────────────────
# DATA VALIDATION COPIER — preserves dropdown accessibility from template
# ─────────────────────────────────────────────────────────────────────────────
def copy_data_validations(src_wb, dst_wb):
    """
    Copy all data validation rules (dropdown lists) from every sheet in src_wb
    to the matching sheet in dst_wb.  openpyxl loses these when it re-saves a
    workbook that was opened after shutil.copy2, so we must re-apply them.
    """
    for sheet_name in src_wb.sheetnames:
        if sheet_name not in dst_wb.sheetnames:
            continue
        src_ws = src_wb[sheet_name]
        dst_ws = dst_wb[sheet_name]

        # Remove any existing (possibly broken) validations on the destination
        dst_ws.data_validations.dataValidation = []

        for dv in src_ws.data_validations.dataValidation:
            new_dv = copy.copy(dv)
            dst_ws.add_data_validation(new_dv)


# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSER  — extracts data from a TSO PDF (4-page format)
# ─────────────────────────────────────────────────────────────────────────────
def parse_pdf(pdf_path):
    data = {'meta': {}, 'bom': [], 'inhouse_rm': {}, 'tool_ops': [], 'assy_ops': [],
            'source': 'pdf'}

    with pdfplumber.open(pdf_path) as pdf:

        # Page 1: meta
        if pdf.pages:
            for row in (pdf.pages[0].extract_tables() or [[]])[0]:
                rc = [cl(c) for c in row]
                if not rc: continue
                if rc[0] == 'Date' and len(rc) >= 2:
                    data['meta']['date'] = rc[1]
                    for i, v in enumerate(rc):
                        if v == 'Project Name' and i+1 < len(rc):
                            data['meta']['project'] = rc[i+1]
                if rc[0] == 'Supplier Name' and len(rc) >= 2:
                    data['meta']['supplier'] = rc[1]
                    for i, v in enumerate(rc):
                        if 'stamping' in str(v).lower() and i+1 < len(rc):
                            data['meta']['stamping_loc'] = rc[i+1]
                if any('end items' in str(c).lower() for c in rc if c):
                    data['meta']['end_items'] = rc[1] if len(rc) > 1 else ''
                    for i, v in enumerate(rc):
                        if 'Welding' in str(v) and i+1 < len(rc):
                            data['meta']['welding_loc'] = rc[i+1]

        # Page 2: BOM
        if len(pdf.pages) >= 2:
            tbls = pdf.pages[1].extract_tables()
            if tbls:
                for row in tbls[0][1:]:
                    rc = [cl(c) for c in row]
                    if not rc[0]: continue
                    surface_treatment = rc[13] if len(rc) > 13 else ''
                    data['bom'].append({
                        'sno':               rc[0],
                        'part_no':           rc[1]  if len(rc) > 1  else '',
                        'type_part':         rc[5]  if len(rc) > 5  else '',
                        'cad_wt':            rc[7]  if len(rc) > 7  else '',
                        'material':          rc[8]  if len(rc) > 8  else '',
                        'thickness':         rc[9]  if len(rc) > 9  else '',
                        'qty_assy':          rc[10] if len(rc) > 10 else '',
                        'qty_veh':           rc[11] if len(rc) > 11 else '',
                        'surface_treatment': 'Yes' if surface_treatment and surface_treatment not in ('', '---') else 'No',
                    })

        # Page 3: RM + tool ops
        if len(pdf.pages) >= 3:
            tbls = pdf.pages[2].extract_tables()
            if not tbls: return data
            tbl = tbls[0]
            main_idx = None
            for idx, row in enumerate(tbl):
                rc = [cl(c) for c in row]
                if rc and rc[0] == '1':
                    main_idx = idx
                    data['inhouse_rm'] = {
                        'input_wt':   rc[32] if len(rc) > 32 else '',
                        'output_wt':  rc[33] if len(rc) > 33 else '',
                        'blank_thk':  rc[17] if len(rc) > 17 else '',
                        'rm_grade':   rc[10] if len(rc) > 10 else '',
                        'sheet_l':    rc[24] if len(rc) > 24 else '',
                        'sheet_w':    rc[23] if len(rc) > 23 else '',
                        'rm_supplier': rc[36] if len(rc) > 36 else '',
                    }
                    break
            if main_idx is None: return data

            def extract_op(rc, n1, n2, lc):
                if len(rc) <= lc + 6: return None
                p1 = rc[n1] if n1 < len(rc) else ''
                p2 = rc[n2] if n2 < len(rc) else ''
                name = f"{p1} {p2}".strip() if p2 and p2.upper() == 'TOOL' else (f"{p1} {p2}".strip() if p2 else p1)
                name = name.strip()
                if not name: return None
                u = name.upper()
                if not any(k in u for k in ['BLANK','FORM','PIERC','PUNCH','INSPECT','SHEAR','PANEL','WELD','TOOL','TAP']):
                    return None
                return {
                    'raw_name':  name,
                    'tool_l':    rc[lc]   if len(rc) > lc   else '',
                    'tool_w':    rc[lc+1] if len(rc) > lc+1 else '',
                    'tool_h':    rc[lc+2] if len(rc) > lc+2 else '',
                    'tonnage':   rc[lc+3] if len(rc) > lc+3 else '',
                    'press':     rc[lc+4] if len(rc) > lc+4 else '',
                    'parts_per': rc[lc+5] if len(rc) > lc+5 else '',
                    'construct': rc[lc+6] if len(rc) > lc+6 else '',
                }

            op = extract_op([cl(c) for c in tbl[main_idx]], 40, 41, 42)
            if op: data['tool_ops'].append(op)
            for row in tbl[main_idx + 1:]:
                rc = [cl(c) for c in row]
                if not any(v for v in rc): continue
                op = extract_op(rc, 40, 41, 42)
                if op: data['tool_ops'].append(op)

        # Page 4: assembly ops
        if len(pdf.pages) >= 4:
            tbls = pdf.pages[3].extract_tables()
            if tbls:
                tbl4 = tbls[0]
                current_side = ''
                for row in tbl4:
                    rc = [str(c).strip() if c else '' for c in row]
                    if not any(v for v in rc): continue
                    if rc[0] in ('1','2') and len(rc) > 3 and rc[3]:
                        pup = rc[3].upper()
                        if ' RH' in pup or pup.endswith('RH'): current_side = 'RH'
                        elif ' LH' in pup or pup.endswith('LH'): current_side = 'LH'
                    ftg_raw      = rc[25] if len(rc) > 25 else ''
                    ftg_type_raw = rc[26] if len(rc) > 26 else ''
                    if not ftg_raw or ftg_raw in ('---', 'FTG Description'):
                        continue
                    ftg_up = ftg_raw.upper()
                    is_orbital = 'ORBIT' in ftg_up
                    is_assy    = (not is_orbital) and any(
                        k in ftg_up for k in ['CHECK','ASSY','INSPECT','FIXTURE'])
                    if not (is_orbital or is_assy):
                        continue
                    side_sfx = (' ' + current_side) if current_side else ''
                    if is_orbital:
                        clean_desc  = 'Orbital Riveting Fixture' + side_sfx
                        sub_op_name = 'Orbital Rivetting'
                    else:
                        clean_desc  = ftg_raw.strip() + side_sfx
                        sub_op_name = 'Assy inspection'
                    data['assy_ops'].append({
                        'ftg_desc':    clean_desc,
                        'ftg_type':    ftg_type_raw,
                        'sub_op_name': sub_op_name,
                        'is_orbital':  is_orbital,
                        'side':        current_side,
                    })

    return data


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSER  — extracts data from a filled TSO Excel workbook
# ─────────────────────────────────────────────────────────────────────────────
def parse_excel(excel_path):
    data = {
        'meta': {},
        'bom': [],
        'inhouse_rm': {},
        'tool_ops': [],
        'assy_ops': [],
        'process_rows': [],
        'source': 'excel'
    }

    wb = load_workbook(str(excel_path), data_only=True)

    if 'TSO Summary' in wb.sheetnames:
        ws = wb['TSO Summary']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            key = str(row[0]).strip().lower()
            val = cl(row[1]) if len(row) > 1 else ''
            if 'date' in key and 'sign' not in key: data['meta']['date'] = val
            elif 'project' in key:                  data['meta']['project'] = val
            elif 'supplier' in key and 'sign' not in key: data['meta']['supplier'] = val
            elif 'stamping' in key:                 data['meta']['stamping_loc'] = val
            elif 'welding' in key:                  data['meta']['welding_loc'] = val
            elif 'end items' in key:                data['meta']['end_items'] = val

    if 'BOM Template' in wb.sheetnames:
        ws = wb['BOM Template']
        for r in range(3, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v for v in row): continue
            level    = cl(row[0])
            part_no  = cl(row[1])
            if not part_no: continue
            data['bom'].append({
                'sno':               '1' if level == '0' else '1.1',
                'part_no':           part_no,
                'type_part':         cl(row[5]) if len(row) > 5 else '',
                'cad_wt':            cl(row[11]) if len(row) > 11 else '',
                'material':          '',
                'thickness':         '',
                'qty_assy':          cl(row[10]) if len(row) > 10 else '',
                'qty_veh':           cl(row[10]) if len(row) > 10 else '',
                'surface_treatment': cl(row[13]) if len(row) > 13 else '',
                '_level':            level,
            })

    if 'Inhouse RM' in wb.sheetnames:
        ws = wb['Inhouse RM']
        if ws.max_row >= 3:
            r3 = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
            data['inhouse_rm'] = {
                'input_wt':  cl(r3[15]) if len(r3) > 15 else '',
                'output_wt': cl(r3[17]) if len(r3) > 17 else '',
                'blank_thk': cl(r3[12]) if len(r3) > 12 else '',
                'rm_grade':  cl(r3[4])  if len(r3) > 4  else '',
            }
            if len(r3) > 5 and r3[5]:
                child_pno = cl(r3[1]) if len(r3) > 1 else ''
                for p in data['bom']:
                    if p['part_no'] == child_pno or child_pno in p['part_no']:
                        p['material'] = cl(r3[5])
                        p['thickness'] = cl(r3[12]) if len(r3) > 12 else ''
                        break

    if 'Inhouse Process' in wb.sheetnames:
        ws = wb['Inhouse Process']
        for r in range(3, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v for v in row): continue
            data['process_rows'].append({
                'level': cl(row[0]) if len(row) > 0 else '',
                'part_no': cl(row[1]) if len(row) > 1 else '',
                'desc': cl(row[2]) if len(row) > 2 else '',
                'mfg': cl(row[3]) if len(row) > 3 else '',
                'sub_op': cl(row[4]) if len(row) > 4 else '',
                'sub_op_name': cl(row[5]) if len(row) > 5 else '',
                'op_no': row[6] if len(row) > 6 else '',
                'ftg': cl(row[7]) if len(row) > 7 else '',
                'ftg_name': cl(row[8]) if len(row) > 8 else '',
                'ftg_qty': row[9] if len(row) > 9 else '',
                'mach_make': cl(row[10]) if len(row) > 10 else '',
                'mach_spec': cl(row[11]) if len(row) > 11 else '',
                'p1_type': cl(row[12]) if len(row) > 12 else '',
                'p1_uom': cl(row[13]) if len(row) > 13 else '',
                'p1_val': row[14] if len(row) > 14 else '',
                'p2_type': cl(row[15]) if len(row) > 15 else '',
                'p2_uom': cl(row[16]) if len(row) > 16 else '',
                'p2_val': row[17] if len(row) > 17 else '',
                'remarks': cl(row[18]) if len(row) > 18 else '',
            })
            mfg      = cl(row[3]) if len(row) > 3 else ''
            sub_op   = cl(row[4]) if len(row) > 4 else ''
            ftg_name = cl(row[8]) if len(row) > 8 else ''
            tonnage  = cl(row[17]) if len(row) > 17 else ''
            construct = cl(row[18]) if len(row) > 18 else ''
            if not mfg and not sub_op: continue
            raw_name = ftg_name or sub_op
            if raw_name:
                data['tool_ops'].append({
                    'raw_name':  raw_name.upper(),
                    'tool_l':    '',
                    'tool_w':    '',
                    'tool_h':    '',
                    'tonnage':   tonnage,
                    'press':     cl(row[11]) if len(row) > 11 else '',
                    'parts_per': '',
                    'construct': construct,
                    '_mfg':      mfg,
                    '_sub_op':   sub_op,
                    '_ftg_name': ftg_name,
                })

    return data


# ─────────────────────────────────────────────────────────────────────────────
# UNIFIED PARSER  — detects file type and routes accordingly
# ─────────────────────────────────────────────────────────────────────────────
def parse_input(input_path):
    """Parse PDF or Excel input. Returns unified data dict."""
    suffix = Path(str(input_path)).suffix.lower()
    if suffix == '.pdf':
        return parse_pdf(input_path)
    elif suffix in ('.xlsx', '.xls', '.xlsm'):
        return parse_excel(input_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Use .pdf or .xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

# Machine Make hardcoded value for all stamping ops (Blanking, Forming, Piercing)
MACHINE_MAKE_STAMPING = 'IGSEC'

def write_excel(data, template_path, out_path):
    flags = []
    shutil.copy2(str(template_path), str(out_path))

    # Load template separately (read-only) so we can copy its data validations later
    template_wb = load_workbook(str(template_path))

    wb = load_workbook(str(out_path))
    lib = load_library(wb)

    meta         = data['meta']
    bom          = data['bom']
    stamp        = data['inhouse_rm']
    tool_ops     = data['tool_ops']
    process_rows = data.get('process_rows', [])
    source       = data.get('source', 'pdf')

    supplier_full = meta.get('supplier', '')
    supplier_short = re.sub(
        r'\s+(PVT\.?\s*LTD\.?|LTD\.?|INC\.?|CORP\.?).*$', '',
        supplier_full, flags=re.IGNORECASE
    ).strip()
    if supplier_short:
        supplier_short = supplier_short.split()[0].title()
    city = meta.get('stamping_loc', '').strip().title()

    pdf_bom = {}
    for p in bom:
        pdf_bom[str(p['part_no']).strip()] = p

    bom_inhouse_vals  = lib.get('D', [])
    bom_import_vals   = lib.get('E', [])
    bom_country_vals  = lib.get('F', [])
    bom_sf_vals       = lib.get('G', [])
    bom_st_vals       = lib.get('H', [])
    bom_ht_vals       = lib.get('I', [])

    def bom_yn(val_list, want):
        w = want.lower()
        for v in val_list:
            if v.lower() == w:
                return v
        return want

    # ── BOM Template ──────────────────────────────────────────────────────
    ws_bom = wb['BOM Template']

    ws_bom.insert_rows(2)
    from openpyxl import load_workbook as _lw
    _tpl_check = [ws_bom.cell(3, c).value for c in range(1,3)]
    if _tpl_check[0] is not None:
        sv(ws_bom, 2, 1, _tpl_check[0])
        sv(ws_bom, 2, 2, _tpl_check[1])
        sv(ws_bom, 2, 3, ws_bom.cell(3,3).value)

    for r in range(3, ws_bom.max_row + 1):
        pno   = str(ws_bom.cell(r, 2).value or '').strip()
        level = str(ws_bom.cell(r, 1).value or '').strip()
        if not pno:
            continue

        pdf_p = pdf_bom.get(pno, {})
        type_part = str(pdf_p.get('type_part', '')).upper()

        if not ws_bom.cell(r, 4).value:
            sv(ws_bom, r, 4, bom_yn(lib.get('C',[]), 'No'))

        if not ws_bom.cell(r, 5).value:
            sv(ws_bom, r, 5, bom_yn(lib.get('C',[]), 'No'))

        if not ws_bom.cell(r, 6).value:
            if type_part in ('BOU', 'BOP', 'BOP-CONSIGNEE', 'BOP-DIRECTED'):
                val = find_in_lib(['bop'], bom_inhouse_vals) or 'BOP'
            else:
                val = find_in_lib(['inhouse'], bom_inhouse_vals) or 'Inhouse'
            sv(ws_bom, r, 6, val)

        if not ws_bom.cell(r, 7).value and supplier_short:
            sv(ws_bom, r, 7, supplier_short)

        if not ws_bom.cell(r, 8).value:
            val = find_in_lib(['local'], bom_import_vals)
            sv(ws_bom, r, 8, val)

        if not ws_bom.cell(r, 9).value and city:
            sv(ws_bom, r, 9, city)

        if not ws_bom.cell(r, 10).value:
            val = find_in_lib(['india'], bom_country_vals)
            sv(ws_bom, r, 10, val)

        if not ws_bom.cell(r, 12).value and pdf_p.get('cad_wt'):
            try:    sv(ws_bom, r, 12, float(pdf_p['cad_wt']))
            except: sv(ws_bom, r, 12, pdf_p['cad_wt'])

        if not ws_bom.cell(r, 13).value:
            sv(ws_bom, r, 13, bom_yn(bom_sf_vals, 'No'))

        if not ws_bom.cell(r, 14).value:
            st_val = pdf_p.get('surface_treatment', 'No')
            sv(ws_bom, r, 14, bom_yn(bom_st_vals, st_val))

        if not ws_bom.cell(r, 15).value:
            sv(ws_bom, r, 15, bom_yn(bom_ht_vals, 'No'))

    # ── Inhouse RM ────────────────────────────────────────────────────────
    ws_rm = wb['Inhouse RM']

    child_part = None
    for p in bom:
        sno = str(p.get('sno', '')).strip()
        tp  = str(p.get('type_part', '')).upper()
        if '.' in sno and tp not in ('BOU', 'BOP'):
            child_part = p
            break
    if child_part is None:
        child_part = next((p for p in bom if '.' in str(p.get('sno',''))), None)

    rm_grade_vals   = lib.get('AC', [])
    rm_country_vals = lib.get('AD', [])
    rm_param_vals   = lib.get('AE', [])
    rm_uom_vals     = lib.get('AF', [])

    r = 3

    if not ws_rm.cell(r, 1).value:
        sv(ws_rm, r, 1, '1')

    if not ws_rm.cell(r, 2).value and child_part:
        sv(ws_rm, r, 2, child_part['part_no'])

    if not ws_rm.cell(r, 3).value and child_part:
        desc = ''
        for bom_r in range(2, ws_bom.max_row + 1):
            if str(ws_bom.cell(bom_r, 2).value or '').strip() == str(child_part['part_no']).strip():
                desc = str(ws_bom.cell(bom_r, 3).value or '').strip()
                break
        sv(ws_rm, r, 3, desc or child_part.get('desc', ''))

    if not ws_rm.cell(r, 5).value:
        rm_grade_raw = (child_part.get('material', '') if child_part else '') or stamp.get('rm_grade', '')
        if rm_grade_raw:
            grade_match = None
            for g in rm_grade_vals:
                def ng(s): return re.sub(r'[\s\-]','',s).upper()
                if ng(rm_grade_raw[:10]) in ng(g):
                    grade_match = g
                    break
            if grade_match:
                sv(ws_rm, r, 5, grade_match)
            else:
                norm = re.sub(r'G\s*0+(\d{2})(\d{4})', r'G-00-\2', rm_grade_raw)
                norm = re.sub(r'\bMM\s*(\d)', r'MM \1', norm).strip()
                sv(ws_rm, r, 6, norm)
                flags.append(
                    f"⚠ Inhouse RM — RM Grade '{norm}' not found in Library dropdown. "
                    f"Written to col F (Others). Please select correct grade in col E."
                )

    if not ws_rm.cell(r, 7).value:
        flags.append('⚠ Inhouse RM — Raw Material Source (col G): Not available in PDF. Please fill manually.')

    if not ws_rm.cell(r, 8).value:
        val = find_in_lib(['india'], rm_country_vals)
        sv(ws_rm, r, 8, val)

    if not ws_rm.cell(r, 9).value:
        val = find_in_lib(['weight'], rm_param_vals)
        sv(ws_rm, r, 9, val)

    if not ws_rm.cell(r, 10).value:
        val = find_in_lib(['kg'], rm_uom_vals)
        sv(ws_rm, r, 10, val)

    if not ws_rm.cell(r, 11).value:
        sheet_l = stamp.get('sheet_l', '')
        if sheet_l:
            try:    sv(ws_rm, r, 11, round(float(sheet_l) / 1000, 4))
            except: pass

    if not ws_rm.cell(r, 12).value:
        sheet_w = stamp.get('sheet_w', '')
        if sheet_w:
            try:    sv(ws_rm, r, 12, round(float(sheet_w) / 1000, 4))
            except: pass

    if not ws_rm.cell(r, 13).value:
        blank_thk = stamp.get('blank_thk', '')
        if blank_thk:
            try:    sv(ws_rm, r, 13, float(blank_thk))
            except: sv(ws_rm, r, 13, blank_thk)

    raw_g = stamp.get('input_wt', '')
    if raw_g and not ws_rm.cell(r, 16).value:
        try:    sv(ws_rm, r, 16, float(raw_g))
        except: sv(ws_rm, r, 16, raw_g)

    raw_n = stamp.get('output_wt', '')
    cad_net = child_part.get('cad_wt', '') if child_part else ''
    try:
        best_net = float(cad_net) if cad_net else (float(raw_n) if raw_n else None)
        if best_net is not None:
            sv(ws_rm, r, 18, best_net)
    except:
        if raw_n: sv(ws_rm, r, 18, raw_n)

    ws_rm.cell(row=r, column=17, value='=P3-R3')
    ws_rm.cell(row=r, column=19, value='=R3/P3*100')

    # ── Inhouse Process ───────────────────────────────────────────────────
    ws_proc = wb['Inhouse Process']

    if source == 'excel' and process_rows:
        for idx, prow in enumerate(process_rows, start=3):
            sv(ws_proc, idx, 1,  prow.get('level'))
            sv(ws_proc, idx, 2,  prow.get('part_no'))
            sv(ws_proc, idx, 3,  prow.get('desc'))
            sv(ws_proc, idx, 4,  prow.get('mfg'))
            sv(ws_proc, idx, 5,  prow.get('sub_op'))
            sv(ws_proc, idx, 6,  prow.get('sub_op_name'))
            sv(ws_proc, idx, 7,  prow.get('op_no'))
            sv(ws_proc, idx, 8,  prow.get('ftg'))
            sv(ws_proc, idx, 9,  prow.get('ftg_name'))
            sv(ws_proc, idx, 10, prow.get('ftg_qty'))
            sv(ws_proc, idx, 11, prow.get('mach_make'))
            sv(ws_proc, idx, 12, prow.get('mach_spec'))
            sv(ws_proc, idx, 13, prow.get('p1_type'))
            sv(ws_proc, idx, 14, prow.get('p1_uom'))
            sv(ws_proc, idx, 15, prow.get('p1_val'))
            sv(ws_proc, idx, 16, prow.get('p2_type'))
            sv(ws_proc, idx, 17, prow.get('p2_uom'))
            sv(ws_proc, idx, 18, prow.get('p2_val'))
            sv(ws_proc, idx, 19, prow.get('remarks'))

        # Copy data validations from template before saving
        copy_data_validations(template_wb, wb)
        wb.save(str(out_path))
        return flags

    # ── Library lookups for Inhouse Process ───────────────────────────────
    mfg_vals  = lib.get('V', [])
    sub_vals  = lib.get('W', [])
    ftg_vals  = lib.get('X', [])
    p1t_vals  = lib.get('Y', [])
    p1u_vals  = lib.get('Z', [])
    p2t_vals  = lib.get('AA', [])
    p2u_vals  = lib.get('AB', [])

    mfg_others     = find_in_lib(['others'],               mfg_vals)
    mfg_sheetmetal = find_in_lib(['sheet metal', 'cold'],  mfg_vals)
    ftg_tool_mm    = find_in_lib(['tool', 'm&m'],          ftg_vals)
    ftg_gauge_mm   = find_in_lib(['gauge', 'm&m'],         ftg_vals)
    ftg_fix_mm     = find_in_lib(['fixture', 'm&m'],       ftg_vals)
    p1t_strokes    = find_in_lib(['strokes'],              p1t_vals)
    p1t_weight     = next((v for v in p1t_vals if v.strip().lower() == 'weight'), None) or find_in_lib(['weight'], p1t_vals)
    p1t_pstroke    = find_in_lib(['parts', 'stroke'],      p1t_vals)
    p1t_pieces     = find_in_lib(['pieces'],               p1t_vals)
    p1t_others     = find_in_lib(['others'],               p1t_vals)
    p1u_nos        = find_in_lib(['nos'],                  p1u_vals)
    p1u_kgs        = find_in_lib(['kgs'],                  p1u_vals)
    p2t_tonnage    = find_in_lib(['tonnage'],              p2t_vals)
    p2u_others     = find_in_lib(['others'],               p2u_vals)
    sub_rivet      = find_in_lib(['riveting'],             sub_vals)
    sub_inspect    = find_in_lib(['inspection'],           sub_vals)
    sub_shear      = find_in_lib(['shearing'],             sub_vals)
    sub_blank      = find_in_lib(['blank', 'pierce'],      sub_vals)
    sub_forming    = find_in_lib(['forming'],              sub_vals)
    sub_pierc1     = find_in_lib(['piercing', '1'],        sub_vals, prefer_no_suffix=False)
    sub_pierc2     = find_in_lib(['piercing', '2'],        sub_vals, prefer_no_suffix=False)
    for sv_name, terms in [('sub_pierc1', ['piercing','1']), ('sub_pierc2', ['piercing','2'])]:
        matches = [v for v in sub_vals if all(t in v.lower() for t in terms)]
        plain   = [m for m in matches if m.lower().startswith('piercing')]
        if plain:
            if sv_name == 'sub_pierc1': sub_pierc1 = plain[0]
            else:                       sub_pierc2 = plain[0]

    cats = {k: None for k in ['BLANK','FORMING','PIERC1','PIERC2','INSPECT']}
    for op in tool_ops:
        u = (op.get('raw_name') or '').upper()
        is_blank   = 'BLANK' in u and 'FINE' not in u and 'PROFILE' not in u
        is_1st     = '1ST' in u or 'FIRST' in u
        is_2nd     = '2ND' in u or 'SECOND' in u
        is_punch   = 'PIERC' in u or 'PUNCH' in u
        is_inspect = 'INSPECT' in u or 'PANEL' in u or 'CHECK' in u

        if   is_blank                      and cats['BLANK']   is None: cats['BLANK']   = op
        elif is_1st and is_punch           and cats['PIERC1']  is None: cats['PIERC1']  = op
        elif is_2nd and is_punch           and cats['PIERC2']  is None: cats['PIERC2']  = op
        elif 'CAM' in u and 'PIERC' in u   and cats['PIERC1']  is None: cats['PIERC1']  = op
        elif is_punch and not is_blank     and cats['PIERC2']  is None: cats['PIERC2']  = op
        elif is_inspect                    and cats['INSPECT'] is None: cats['INSPECT'] = op
        elif 'FORM' in u and not is_blank  and cats['FORMING'] is None: cats['FORMING'] = op

    def press_tc(op):
        v = (op or {}).get('press', '')
        return title_case(v) if v and v == v.upper() else v

    def tonnage_v(op): return (op or {}).get('tonnage', '')

    def parts_per_v(op):
        raw = (op or {}).get('parts_per', '')
        try:    return int(raw) if raw and str(raw).strip() else ''
        except: return raw

    def construct_v(op):
        v = (op or {}).get('construct', '')
        if not v: return ''
        tc = title_case(v) if v == v.upper() else v
        FAB = ('fabricat', 'cast', 'weld', 'machined', 'forged')
        return tc if any(k in tc.lower() for k in FAB) else ''

    assy_pno = assy_desc = ''
    for bom_r in range(2, ws_bom.max_row + 1):
        lvl = str(ws_bom.cell(bom_r, 1).value or '').strip()
        if lvl == '0':
            assy_pno  = str(ws_bom.cell(bom_r, 2).value or '').strip()
            assy_desc = str(ws_bom.cell(bom_r, 3).value or '').strip()
            break

    child_pno = child_desc = ''
    for p in bom:
        sno = str(p.get('sno', '')).strip()
        tp  = str(p.get('type_part', '')).upper()
        if '.' in sno and tp not in ('BOU', 'BOP'):
            child_pno = str(p['part_no']).strip()
            for bom_r in range(2, ws_bom.max_row + 1):
                if str(ws_bom.cell(bom_r, 2).value or '').strip() == child_pno:
                    child_desc = str(ws_bom.cell(bom_r, 3).value or '').strip()
                    break
            break

    gross_val = stamp.get('input_wt', '')
    try:    gross = float(gross_val) if gross_val else ''
    except: gross = gross_val

    cur = 3

    assy_ops = data.get('assy_ops', [])
    rivet_aop      = next((a for a in assy_ops if a.get('is_orbital')), None)
    insp_aop       = next((a for a in assy_ops if not a.get('is_orbital') and a.get('ftg_desc')), None)
    rivet_ftg_name = rivet_aop['ftg_desc']    if rivet_aop else ''
    insp_ftg_name  = insp_aop['ftg_desc']     if insp_aop  else ''
    rivet_sub_nm   = rivet_aop['sub_op_name'] if rivet_aop else ''
    insp_sub_nm    = insp_aop['sub_op_name']  if insp_aop  else ''

    # ── R3: Assembly — Riveting ────────────────────────────────────────────
    sv(ws_proc, cur, 1, '0');        sv(ws_proc, cur, 2, assy_pno);  sv(ws_proc, cur, 3, assy_desc)
    sv(ws_proc, cur, 4, mfg_others); sv(ws_proc, cur, 5, sub_rivet)
    if rivet_sub_nm: sv(ws_proc, cur, 6, rivet_sub_nm)
    sv(ws_proc, cur, 7, 10)
    sv(ws_proc, cur, 8, ftg_fix_mm)
    if rivet_ftg_name: sv(ws_proc, cur, 9, rivet_ftg_name)
    sv(ws_proc, cur, 10, 1)
    sv(ws_proc, cur, 13, p1t_strokes); sv(ws_proc, cur, 14, p1u_nos); sv(ws_proc, cur, 15, 1)
    cur += 1

    # ── R4: Assembly — Assy Inspection ────────────────────────────────────
    sv(ws_proc, cur, 4, mfg_others);  sv(ws_proc, cur, 5, sub_inspect)
    if insp_sub_nm: sv(ws_proc, cur, 6, insp_sub_nm)
    sv(ws_proc, cur, 7, 20)
    sv(ws_proc, cur, 8, ftg_gauge_mm)
    if insp_ftg_name: sv(ws_proc, cur, 9, insp_ftg_name)
    sv(ws_proc, cur, 10, 1)
    sv(ws_proc, cur, 13, p1t_pieces);  sv(ws_proc, cur, 14, p1u_nos); sv(ws_proc, cur, 15, 1)
    cur += 1

    # ── R5: Child — Shearing ──────────────────────────────────────────────
    sv(ws_proc, cur, 1, '1');          sv(ws_proc, cur, 2, child_pno);  sv(ws_proc, cur, 3, child_desc)
    sv(ws_proc, cur, 4, mfg_others);   sv(ws_proc, cur, 5, sub_shear);  sv(ws_proc, cur, 7, 10)
    sv(ws_proc, cur, 13, p1t_weight);  sv(ws_proc, cur, 14, p1u_kgs);   sv(ws_proc, cur, 15, gross)
    cur += 1

    # ── R6+: Stamping ops ─────────────────────────────────────────────────
    # col 11 = Machine Make → hardcoded to IGSEC for all stamping ops
    # col 12 = Machine Spec (press type from PDF)
    stamping_ops = [
        ('BLANK',   sub_blank,   20),
        ('FORMING', sub_forming, 30),
        ('PIERC1',  sub_pierc1,  40),
        ('PIERC2',  sub_pierc2,  50),
        ('INSPECT', sub_inspect, 60),
    ]
    for cat_key, sub_val, op_num in stamping_ops:
        op = cats.get(cat_key)
        if op is None:
            continue

        ftg_name = normalise_ftg_name(op.get('raw_name', ''))
        pp       = parts_per_v(op)
        ton      = tonnage_v(op)
        con      = construct_v(op)
        prs      = press_tc(op)

        if cat_key == 'INSPECT':
            sv(ws_proc, cur, 4, mfg_others)
            sv(ws_proc, cur, 5, sub_val);      sv(ws_proc, cur, 7, op_num)
            sv(ws_proc, cur, 8, ftg_gauge_mm); sv(ws_proc, cur, 9, ftg_name); sv(ws_proc, cur, 10, 1 if ftg_name else '')
            sv(ws_proc, cur, 13, p1t_others);  sv(ws_proc, cur, 14, p1u_nos); sv(ws_proc, cur, 15, 1)
        else:
            sv(ws_proc, cur, 4, mfg_sheetmetal)
            sv(ws_proc, cur, 5, sub_val);      sv(ws_proc, cur, 7, op_num)
            sv(ws_proc, cur, 8, ftg_tool_mm);  sv(ws_proc, cur, 9, ftg_name); sv(ws_proc, cur, 10, 1 if ftg_name else '')
            # ── Machine Make: hardcoded to IGSEC for Blank / Forming / Piercing ──
            sv(ws_proc, cur, 11, MACHINE_MAKE_STAMPING)
            # Machine Spec: press type from PDF
            sv(ws_proc, cur, 12, prs)
            sv(ws_proc, cur, 13, p1t_pstroke); sv(ws_proc, cur, 14, p1u_nos)
            sv(ws_proc, cur, 15, pp if pp != '' else '')
            if ton:
                sv(ws_proc, cur, 16, p2t_tonnage); sv(ws_proc, cur, 17, p2u_others); sv(ws_proc, cur, 18, ton)
            if con:
                sv(ws_proc, cur, 19, con)
        cur += 1

    # ── Restore data validations from template (fixes dropdown accessibility) ──
    copy_data_validations(template_wb, wb)

    wb.save(str(out_path))
    return flags



"""
TSO Converter — Streamlit Web App
==================================
Upload a TSO PDF or Excel + the M&M TSO Download template.
Downloads the populated output Excel instantly.

Run: streamlit run app.py
"""

import streamlit as st
import tempfile, shutil
from pathlib import Path


st.set_page_config(
    page_title="TSO Converter",
    page_icon="📋",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .block-container { padding-top: 2rem; max-width: 760px; }
    .stAlert { border-radius: 8px; }
    div[data-testid="stFileUploader"] { border-radius: 8px; }
    .flag-box {
        background: #FAEEDA; border-left: 3px solid #BA7517;
        padding: 10px 14px; border-radius: 4px;
        font-size: 13px; font-family: monospace; color: #412402;
        margin-bottom: 6px; white-space: pre-wrap; word-break: break-word;
    }
</style>
""", unsafe_allow_html=True)

st.title("TSO Converter")
st.caption("Upload a TSO source file (PDF or Excel) + the M&M TSO Download template → get a populated Excel ready for upload.")

st.divider()

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. TSO source file")
    input_file = st.file_uploader(
        "PDF or Excel",
        type=["pdf", "xlsx"],
        help="The TSO document — either the PDF received from supplier or a previously filled Excel.",
        label_visibility="collapsed",
    )
    if input_file:
        ext = Path(input_file.name).suffix.upper()
        st.success(f"{ext} uploaded — **{input_file.name}**")

with col2:
    st.subheader("2. TSO template Excel")
    template_file = st.file_uploader(
        "M&M TSO Download template (.xlsx)",
        type=["xlsx"],
        help="The blank M&M TSO Download template Excel — must contain the Library sheet with all dropdowns.",
        label_visibility="collapsed",
    )
    if template_file:
        st.success(f"XLSX uploaded — **{template_file.name}**")

st.divider()

if not input_file or not template_file:
    st.info("Upload both files above to enable conversion.", icon="ℹ️")
    st.stop()

if st.button("Convert to Excel", type="primary", use_container_width=True):
    with st.spinner("Reading input and writing Excel…"):
        try:
            with tempfile.TemporaryDirectory() as tmp:
                tmp = Path(tmp)

                input_path    = tmp / input_file.name
                template_path = tmp / template_file.name
                out_name      = Path(input_file.name).stem + "_TSO_output.xlsx"
                out_path      = tmp / out_name

                input_path.write_bytes(input_file.getvalue())
                template_path.write_bytes(template_file.getvalue())

                data  = parse_input(input_path)
                flags = write_excel(data, template_path, out_path)

                output_bytes = out_path.read_bytes()

            st.success("Conversion complete!", icon="✅")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Source",    data.get('source','').upper())
            m2.metric("Project",   data['meta'].get('project', '—'))
            m3.metric("BOM parts", len(data['bom']))
            m4.metric("Tool ops",  len(data['tool_ops']))

            st.download_button(
                label="⬇ Download output Excel",
                data=output_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

            if flags:
                st.warning(f"{len(flags)} field(s) need manual review after download:", icon="⚠️")
                for flag in flags:
                    st.markdown(f'<div class="flag-box">{flag.strip()}</div>', unsafe_allow_html=True)
            else:
                st.info("All fields matched from Library dropdowns — no manual review needed.", icon="✅")

        except Exception as e:
            import traceback
            st.error(f"Conversion failed: {e}", icon="❌")
            with st.expander("Error details"):
                st.code(traceback.format_exc())

st.divider()
st.caption("TSO Converter v4 · Supports PDF and Excel input · All dropdowns sourced from Library sheet at runtime")
